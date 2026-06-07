"""岗位清单 Excel 导出模块。"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable


COLUMNS = [
    "company",
    "group",
    "title",
    "location",
    "deadline",
    "major_requirement",
    "education_requirement",
    "match_score",
    "recommended",
    "priority",
    "recommend_reason",
    "resume_tips",
    "interview_tips",
    "url",
]

HEADERS = [
    "单位",
    "上级集团",
    "岗位",
    "地点",
    "报名时间",
    "专业要求",
    "学历要求",
    "匹配评分",
    "是否推荐",
    "投递优先级",
    "推荐理由",
    "简历优化建议",
    "面试准备建议",
    "报名链接",
]


def export_jobs(jobs: Iterable[dict], output_path: str = "jobs.xlsx") -> str:
    """Export jobs to xlsx when openpyxl exists; otherwise write csv fallback."""

    jobs = list(jobs)
    path = Path(output_path)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        csv_path = path.with_suffix(".csv")
        with csv_path.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.writer(file)
            writer.writerow(HEADERS)
            for job in jobs:
                writer.writerow([job.get(column, "") for column in COLUMNS])
        return str(csv_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "岗位清单"
    sheet.append(HEADERS)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for job in jobs:
        sheet.append([job.get(column, "") for column in COLUMNS])

    widths = [16, 18, 24, 14, 18, 24, 14, 10, 12, 12, 36, 36, 36, 36]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    workbook.save(path)
    return str(path)


if __name__ == "__main__":
    sample_jobs = [
        {
            "company": "国家电网",
            "group": "国家电网有限公司",
            "title": "信息通信运维岗",
            "location": "四川",
            "deadline": "待确认",
            "major_requirement": "通信工程、电子信息类",
            "education_requirement": "本科及以上",
            "match_score": 90,
            "recommended": "推荐",
            "priority": "高",
            "recommend_reason": "专业和供电所实习经历匹配",
            "resume_tips": "突出通信运维和供电所实习",
            "interview_tips": "准备电力通信和计算机网络基础",
            "url": "https://zhaopin.sgcc.com.cn/",
        }
    ]
    print(export_jobs(sample_jobs))
